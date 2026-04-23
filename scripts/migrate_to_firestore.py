"""
migrate_to_firestore.py
Migra DatabaseInicial.xlsx a Cloud Firestore para GeoViewer AMVA.

Colecciones y filas esperadas:
  Sondeos           198   Proyectos          31
  Estratigrafias  1.409   SPT             1.237
  Limites_Atterberg 1.111  DownHole          775
  Col_Resonante     563   Triaxial_Ciclico  701
  Lin_Sismicas      348   Gamma             392
  Consolidacion      14   Corte_Directo       5
  ─────────────────────────────────────────────
  TOTAL           6.784 documentos

USO:
  py scripts/migrate_to_firestore.py
  py scripts/migrate_to_firestore.py --solo Sondeos SPT   # solo algunas
  py scripts/migrate_to_firestore.py --borrar             # borra antes de subir
"""

import sys, math, argparse
from pathlib import Path
from datetime import datetime, date

import openpyxl
import firebase_admin
from firebase_admin import credentials, firestore

# ── Rutas ────────────────────────────────────────────────────────────────────
ROOT       = Path(__file__).parent.parent
EXCEL_PATH = Path(r"C:\Users\User\OneDrive\Desktop\DatabaseInicial.xlsx")
# Busca automáticamente el archivo adminsdk en la raíz
KEY_MATCHES = list(ROOT.glob("*adminsdk*.json"))
KEY_PATH    = KEY_MATCHES[0] if KEY_MATCHES else ROOT / "serviceAccountKey.json"

BATCH_SIZE = 400   # Firestore permite max 500 operaciones por batch

# ── Esquema: columnas numéricas por hoja ─────────────────────────────────────
NUMERIC_COLS = {
    "Sondeos":            ["Cota", "Nivel_freatico", "Coordenada_Este", "Coordenada_Norte"],
    "Estratigrafias":     ["NumeroCapa", "Espesor"],
    "SPT":                ["Profundidad", "Valor"],
    "Triaxial_Ciclico":   ["Profundidad", "Numero_Ciclico", "Beta", "Epsilon",
                           "Modulo_Elastico", "Modulo_Poisson", "G", "Y"],
    "Consolidacion":      ["Profundidad", "e_value", "Gs", "Cc", "Cs"],
    "Corte_Directo":      ["Profundidad", "Angulo_Friccion", "Cohesion"],
    "Limites_Atterberg":  ["Profundidad", "W", "LL", "LP"],
    "DownHole":           ["Profundidad", "Vp", "Vs"],
    "Lin_Sismicas":       ["Coordenada_Norte_Ini", "Coordenada_Este_Ini",
                           "Coordenada_Norte_Fin", "Coordenada_Este_Fin",
                           "Coordenada_Norte_Centro", "Coordenada_Este_Centro"],
    "Gamma":              ["Profundidad", "Peso_Unitario",
                           "Gravedad_Especifica", "Pensiometro"],
    "Col_Resonante":      ["Profundidad", "Numero_Muestra", "Y", "G", "Beta"],
}

SHEETS_ORDER = [
    "Proyectos",
    "Sondeos",
    "Estratigrafias",
    "SPT",
    "Limites_Atterberg",
    "DownHole",
    "Col_Resonante",
    "Triaxial_Ciclico",
    "Lin_Sismicas",
    "Gamma",
    "Consolidacion",
    "Corte_Directo",
]

# ── Limpieza de valores ───────────────────────────────────────────────────────
def clean(v):
    """Convierte cualquier valor Python a tipo compatible con Firestore."""
    if v is None:
        return None
    if isinstance(v, float) and math.isnan(v):
        return None
    if isinstance(v, (datetime, date)):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, str):
        stripped = v.strip()
        if stripped.upper() == "NULL" or stripped == "":
            return None
        return stripped
    if isinstance(v, (int, float, bool)):
        return v
    return str(v)

def fix_col_name(name: str) -> str:
    """Normaliza nombres de columna con caracteres especiales."""
    if name is None:
        return "col_sin_nombre"
    return (name
            .replace("Específica", "Especifica")
            .replace("é", "e")
            .replace("í", "i")
            .replace("ó", "o")
            .replace("á", "a")
            .replace("ú", "u")
            .replace("ñ", "n")
            .strip())

# ── Lector de Excel ───────────────────────────────────────────────────────────
def read_sheet(wb: openpyxl.Workbook, sheet_name: str) -> list[dict]:
    """Lee una hoja y devuelve lista de dicts limpios."""
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    # Cabecera con nombres normalizados
    raw_headers = [cell for cell in rows[0]]
    headers = [fix_col_name(h) for h in raw_headers]
    numeric = set(NUMERIC_COLS.get(sheet_name, []))

    docs = []
    for row in rows[1:]:
        doc = {}
        for col, val in zip(headers, row):
            cleaned = clean(val)
            # Forzar numérico en columnas que lo son
            if cleaned is not None and col in numeric:
                try:
                    cleaned = float(cleaned)
                except (ValueError, TypeError):
                    cleaned = None
            doc[col] = cleaned
        # Omitir filas completamente vacías
        if any(v is not None for v in doc.values()):
            docs.append(doc)

    return docs

# ── Subida a Firestore ────────────────────────────────────────────────────────
def upload(db, collection_name: str, docs: list[dict], clear_first: bool):
    col_ref = db.collection(collection_name)

    if clear_first:
        # Borra en batches de 400
        existing = col_ref.limit(400).stream()
        batch = db.batch()
        count = 0
        for doc in existing:
            batch.delete(doc.reference)
            count += 1
        if count:
            batch.commit()

    total   = len(docs)
    batch   = db.batch()
    pending = 0
    done    = 0

    for i, doc in enumerate(docs):
        doc_id  = str(doc.get("Id", i))
        doc_ref = col_ref.document(doc_id)
        batch.set(doc_ref, doc)
        pending += 1
        done    += 1

        if pending >= BATCH_SIZE:
            batch.commit()
            batch   = db.batch()
            pending = 0
            pct = done / total * 100
            print(f"\r    {done:>5}/{total}  [{bar(pct)}] {pct:.0f}%", end="", flush=True)

    if pending:
        batch.commit()

    print(f"\r    {done:>5}/{total}  [{bar(100)}] 100%  OK")

def bar(pct: float, width: int = 20) -> str:
    filled = int(pct / 100 * width)
    return "#" * filled + "-" * (width - filled)

# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(description="Migra DatabaseInicial.xlsx a Firestore")
    parser.add_argument("--solo",   nargs="+", metavar="HOJA",
                        help="Migrar solo estas hojas (p.ej. --solo Sondeos SPT)")
    parser.add_argument("--borrar", action="store_true",
                        help="Borrar documentos existentes antes de subir")
    args = parser.parse_args()

    # Validaciones
    if not EXCEL_PATH.exists():
        print(f"ERROR: No se encontro el Excel en:\n  {EXCEL_PATH}")
        sys.exit(1)
    if not KEY_PATH.exists():
        print(f"ERROR: No se encontro la service account key en:\n  {KEY_PATH}")
        print("Descargala desde Firebase Console > Configuracion > Cuentas de servicio")
        sys.exit(1)

    target_sheets = args.solo if args.solo else SHEETS_ORDER

    print("=" * 56)
    print("  GeoViewer AMVA — Migracion a Firestore")
    print("=" * 56)
    print(f"  Excel : {EXCEL_PATH.name}")
    print(f"  Key   : {KEY_PATH.name}")
    print(f"  Hojas : {', '.join(target_sheets)}")
    if args.borrar:
        print("  Modo  : BORRAR + subir")
    print("=" * 56)

    # Inicializar Firebase
    cred = credentials.Certificate(str(KEY_PATH))
    firebase_admin.initialize_app(cred)
    db = firestore.client()

    # Abrir Excel
    print("\nAbriendo Excel...", end="", flush=True)
    wb = openpyxl.load_workbook(str(EXCEL_PATH), read_only=True, data_only=True)
    print(f" OK  ({len(wb.sheetnames)} hojas)")

    total_docs = 0
    errores    = []

    for sheet in target_sheets:
        if sheet not in wb.sheetnames:
            print(f"\n  [AVISO] Hoja '{sheet}' no encontrada, omitiendo.")
            continue

        print(f"\n  {sheet}")
        try:
            docs = read_sheet(wb, sheet)
            print(f"    Leidos: {len(docs)} documentos")
            upload(db, sheet, docs, clear_first=args.borrar)
            total_docs += len(docs)
        except Exception as e:
            print(f"    ERROR: {e}")
            errores.append((sheet, str(e)))

    print("\n" + "=" * 56)
    print(f"  Completado: {total_docs:,} documentos subidos")
    if errores:
        print(f"  Errores en: {', '.join(e[0] for e in errores)}")
        for sheet, msg in errores:
            print(f"    {sheet}: {msg}")
    else:
        print("  Sin errores")
    print("=" * 56)
    print("\n  Firestore:")
    print("  https://console.firebase.google.com/project/geoviewer-amva/firestore")

if __name__ == "__main__":
    main()
